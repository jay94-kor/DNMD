import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import streamlit as st
from utils import format_currency, sanitize_sheet_title

def generate_summary_excel():
    event_data = st.session_state.event_data
    event_name = event_data.get('event_name', '무제')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    summary_filename = f"이벤트_기획_정의서_{event_name}_{timestamp}.xlsx"

    try:
        create_excel_summary(event_data, summary_filename)
        st.success(f"엑셀 정의서가 성공적으로 생성되었습니다: {summary_filename}")

        with open(summary_filename, "rb") as file:
            st.download_button(label="전체 행사 요약 정의서 다운로드", data=file, file_name=summary_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        for category, component in event_data.get('components', {}).items():
            category_filename = f"발주요청서_{category}_{event_name}_{timestamp}.xlsx"
            create_category_excel(event_data, category, component, category_filename)
            try:
                with open(category_filename, "rb") as file:
                    st.download_button(label=f"{category} 발주요청서 다운로드", data=file, file_name=category_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"download_{category}")
            except FileNotFoundError:
                st.error(f"{category_filename} 파일을 찾을 수 없습니다.")

    except Exception as e:
        st.error(f"엑셀 파일 생성 중 오류가 발생했습니다: {str(e)}")
        st.error("오류 상세 정보:")
        st.exception(e)

def create_excel_summary(event_data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "이벤트 기획 정의서"

    # 제목
    ws['A1'] = "이벤트 기획 정의서"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:G1')

    # 기본 정보
    basic_info = [
        ("프로젝트명", event_data.get('event_name', '')),
        ("클라이언트명", event_data.get('client_name', '')),
        ("담당 PM", event_data.get('manager_name', '')),
        ("담당자 연락처", event_data.get('manager_contact', '')),
    ]

    for row, (key, value) in enumerate(basic_info, start=3):
        ws[f'A{row}'] = key
        ws[f'B{row}'] = value

    # 용역 구성 요소
    row = len(basic_info) + 4
    ws[f'A{row}'] = "용역 구성 요소"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    headers = ["카테고리", "상태", "아이템", "수량", "가격"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)

    for category, component in event_data.get('components', {}).items():
        row += 1
        ws[f'A{row}'] = category
        ws[f'B{row}'] = component.get('status', '')
        for item in component.get('items', []):
            row += 1
            ws[f'C{row}'] = item.get('name', '')
            ws[f'D{row}'] = item.get('quantity', '')
            ws[f'E{row}'] = format_currency(item.get('price', 0))

    wb.save(filename)

def create_category_excel(event_data, category, component, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_title(category)

    # 제목
    ws['A1'] = f"{category} 발주요청서"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')

    # 기본 정보
    basic_info = [
        ("프로젝트명", event_data.get('event_name', '')),
        ("클라이언트명", event_data.get('client_name', '')),
        ("담당 PM", event_data.get('manager_name', '')),
        ("담당자 연락처", event_data.get('manager_contact', '')),
    ]

    for row, (key, value) in enumerate(basic_info, start=3):
        ws[f'A{row}'] = key
        ws[f'B{row}'] = value

    # 아이템 목록
    row = len(basic_info) + 4
    headers = ["아이템", "수량", "가격", "비고"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=header).font = Font(bold=True)

    for item in component.get('items', []):
        row += 1
        ws[f'A{row}'] = item.get('name', '')
        ws[f'B{row}'] = item.get('quantity', '')
        ws[f'C{row}'] = format_currency(item.get('price', 0))

    wb.save(filename)