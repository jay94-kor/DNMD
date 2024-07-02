import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def generate_excel_report(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "이벤트 기획 보고서"

    # 스타일 정의
    title_font = Font(name='맑은 고딕', size=14, bold=True)
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    cell_font = Font(name='맑은 고딕', size=10)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # 제목
    ws['A1'] = "이벤트 기획 보고서"
    ws['A1'].font = title_font
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 기본 정보
    ws['A3'] = "기본 정보"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:G3')

    headers = ["행사명", "클라이언트명", "행사 유형", "규모", "시작일", "종료일", "셋업", "철수"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=header).font = header_font

    values = [data['event_name'], data['client_name'], ', '.join(data['event_type']), data['scale'],
              data['start_date'], data['end_date'], data['setup'], data['teardown']]
    for col, value in enumerate(values, start=1):
        ws.cell(row=5, column=col, value=value).font = cell_font

    # 장소 정보
    ws['A7'] = "장소 정보"
    ws['A7'].font = header_font
    ws['A7'].fill = header_fill
    ws.merge_cells('A7:G7')

    venue_headers = ["장소명", "유형", "주소", "수용 인원", "시설 및 장비"]
    for col, header in enumerate(venue_headers, start=1):
        ws.cell(row=8, column=col, value=header).font = header_font

    venue_values = [data['venue_name'], data['venue_type'], data['address'], data['capacity'], ', '.join(data['facilities'])]
    for col, value in enumerate(venue_values, start=1):
        ws.cell(row=9, column=col, value=value).font = cell_font

    # 예산 정보
    ws['A11'] = "예산 정보"
    ws['A11'].font = header_font
    ws['A11'].fill = header_fill
    ws.merge_cells('A11:G11')

    budget_headers = ["총 예산", "예상 수익률"]
    for col, header in enumerate(budget_headers, start=1):
        ws.cell(row=12, column=col, value=header).font = header_font

    budget_values = [f"{data['total_budget']:,}원", f"{data['expected_profit']:.1f}%"]
    for col, value in enumerate(budget_values, start=1):
        ws.cell(row=13, column=col, value=value).font = cell_font

    # 용역 구성 요소
    ws['A15'] = "용역 구성 요소"
    ws['A15'].font = header_font
    ws['A15'].fill = header_fill
    ws.merge_cells('A15:G15')

    component_headers = ["카테고리", "세부 항목", "상태", "예산", "협력사 컨택 상태", "추가 정보"]
    for col, header in enumerate(component_headers, start=1):
        ws.cell(row=16, column=col, value=header).font = header_font

    for row, component in enumerate(data['service_components'], start=17):
        values = [component['category'], component['subcategory'], component['status'],
                  f"{component['budget']:,}원", component['contact_status'], component['additional_info']]
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value).font = cell_font

    # 열 너비 조정
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15

    return wb

def save_excel_report(data, filename):
    wb = generate_excel_report(data)
    wb.save(filename)