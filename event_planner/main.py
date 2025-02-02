import streamlit as st
from streamlit_option_menu import option_menu
from datetime import date, timedelta, datetime
import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
from typing import Dict, Any, List, Tuple
import logging
import re
from openpyxl.utils.dataframe import dataframe_to_rows
import sqlite3
from contextlib import contextmanager
from functools import lru_cache

# Logging 설정
logging.basicConfig(filename='app.log', level=logging.ERROR)

# JSON 파일 경로 설정
JSON_PATH = os.path.join(os.path.dirname(__file__), 'item_options.json')
CONFIG_PATH = os.path.join(os.path.dirname(__file__), 'config.json')

# 테이블 컬럼 설정
EVENT_TABLE_COLUMNS = [
    'event_name', 'client_name', 'manager_name', 'manager_contact', 'event_type',
    'contract_type', 'scale', 'start_date', 'end_date', 'setup_start', 'teardown',
    'venue_name', 'venue_type', 'address', 'capacity', 'facilities',
    'contract_amount', 'expected_profit', 'components'
]

# EventOptions 클래스 정의
class EventOptions:
    def __init__(self, item_options):
        self.EVENT_TYPES = item_options['EVENT_TYPES']
        self.CONTRACT_TYPES = item_options['CONTRACT_TYPES']
        self.STATUS_OPTIONS = item_options['STATUS_OPTIONS']
        self.MEDIA_ITEMS = item_options['MEDIA_ITEMS']
        self.CATEGORIES = item_options['CATEGORIES']
        self.CATEGORY_ICONS = item_options['CATEGORY_ICONS']

@lru_cache(maxsize=None)
def load_json_file(file_path: str) -> Dict[str, Any]:
    with open(file_path, 'r', encoding='utf-8') as file:
        return json.load(file)

# JSON 파일에서 item_options 로드
item_options = load_json_file(JSON_PATH)

# config.json 파일에서 설정 로드
config = load_json_file(CONFIG_PATH)

event_options = EventOptions(item_options)

# Helper functions
def format_currency(amount: float) -> str:
    return f"{amount:,.0f}"

def format_phone_number(number: str) -> str:
    pattern = r'(\d{3})(\d{3,4})(\d{4})'
    return re.sub(pattern, r'\1-\2-\3', number)

# 단계별 사용자 가이드 추가 함수
def display_guide(guide_text: str) -> None:
    with st.expander("사용자 가이드", expanded=False):
        st.markdown(guide_text)

# 데이터베이스 연결 최적화
@contextmanager
def get_db_connection():
    conn = sqlite3.connect('event_planner.db', check_same_thread=False)
    try:
        yield conn
    finally:
        conn.close()

# 데이터베이스 초기화 함수
def init_db():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            event_data TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        conn.commit()

# JSON 인코더 클래스
class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (date, datetime)):
            return obj.isoformat()
        return super().default(obj)

# 이벤트 데이터 저장 함수
def save_event_data(event_data: Dict[str, Any]) -> None:
    try:
        with get_db_connection() as conn:
            cursor = conn.cursor()
            event_id = event_data.get('id')
            event_data_json = json.dumps(event_data, ensure_ascii=False, cls=CustomJSONEncoder)
            if event_id:
                cursor.execute('''
                UPDATE events SET event_data = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                ''', (event_data_json, event_id))
            else:
                cursor.execute('''
                INSERT INTO events (event_data) VALUES (?)
                ''', (event_data_json,))
                event_data['id'] = cursor.lastrowid
            conn.commit()
    except Exception as e:
        logging.error(f"Error saving event data: {str(e)}")
        logging.error(f"Event data: {event_data}")
        raise

# 이벤트 데이터 로드 함수
@lru_cache(maxsize=32)
def load_event_data(event_id: int) -> Dict[str, Any]:
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT event_data FROM events WHERE id = ?', (event_id,))
        result = cursor.fetchone()
        if result:
            event_data = json.loads(result[0])
            for key, value in event_data.items():
                if isinstance(value, str):
                    try:
                        event_data[key] = datetime.fromisoformat(value).date()
                    except ValueError:
                        pass
            return event_data
        return {}

# 모든 이벤트 가져오기 함수
def get_all_events() -> List[Tuple[int, str, str]]:
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id, event_data, created_at FROM events ORDER BY created_at DESC')
        return [(row[0], json.loads(row[1]).get('event_name', 'Unnamed Event'), row[2]) for row in cursor.fetchall()]

# 앱 시작 시 데이터베이스 초기화
init_db()

# 기본 정보 단계
def basic_info() -> None:
    event_data = st.session_state.event_data
    st.header("기본 정보")

    guide_text = """
    - **용역명**: 프로젝트의 공식 이름을 입력하세요.
    - **클라이언트명**: 고객사의 정확한 법인명을 입력하세요.
    - **담당 PM**: 프로젝트 매니저의 이름을 입력하세요.
    - **담당자 연락처**: 숫자만 입력해주세요 (예: 01012345678).
    """
    display_guide(guide_text)

    handle_general_info(event_data)
    handle_event_type(event_data)
    handle_budget_info(event_data)

    if event_data['event_type'] == "온라인 콘텐츠":
        handle_online_content(event_data)
    elif event_data['event_type'] == "오프라인 이벤트":
        handle_offline_event(event_data)

def handle_general_info(event_data: Dict[str, Any]) -> None:
    st.write(f"현재 예상 참여 관객 수: {event_data.get('scale', 0)}명")

    col1, col2 = st.columns(2)
    with col1:
        event_data['event_name'] = st.text_input("용역명", value=event_data.get('event_name', ''), key="event_name_basic", autocomplete="off")
        event_data['client_name'] = st.text_input("클라이언트명", value=event_data.get('client_name', ''), key="client_name_basic")
    with col2:
        event_data['manager_name'] = st.text_input("담당 PM", value=event_data.get('manager_name', ''), key="manager_name_basic")
        event_data['manager_email'] = st.text_input("담당 PM 이메일", value=event_data.get('manager_email', ''), key="manager_email_basic")

    event_data['manager_position'] = render_option_menu(
        "담당자 직급",
        options=["선임", "책임", "수석"],
        key="manager_position"
    )

    manager_contact = st.text_input(
        "담당자 연락처",
        value=event_data.get('manager_contact', ''),
        help="숫자만 입력해주세요 (예: 01012345678)",
        key="manager_contact_basic"
    )
    if manager_contact:
        manager_contact = ''.join(filter(str.isdigit, manager_contact))
        event_data['manager_contact'] = format_phone_number(manager_contact)

    st.write(f"입력된 연락처: {event_data.get('manager_contact', '')}")

def render_option_menu(label: str, options: List[str], key: str) -> str:
    icons = ["🔹" for _ in options]
    selected = option_menu(
        label, options,
        icons=icons,
        menu_icon="cast",
        default_index=0,
        orientation="horizontal",
        styles={
            "container": {"padding": "5px", "background-color": "#f0f0f0"},
            "icon": {"color": "#ff6347", "font-size": "20px"},
            "nav-link": {"font-size": "18px", "text-align": "center", "margin":"0px", "--hover-color": "#ffcccc", "--icon-color": "#ff6347"},
            "nav-link-selected": {"background-color": "#ff6347", "color": "white", "--icon-color": "white"},
        },
        key=key
    )
    return selected

def handle_event_type(event_data: Dict[str, Any]) -> None:
    col1, col2 = st.columns(2)
    with col1:
        event_data['event_type'] = render_option_menu(
            "용역 유형",
            event_options.EVENT_TYPES,
            "event_type"
        )
    with col2:
        event_data['contract_type'] = render_option_menu(
            "용역 종류",
            event_options.CONTRACT_TYPES,
            "contract_type"
        )

    # 온라인 이벤트 설정
    if event_data['event_type'] == "온라인 콘텐츠":
        event_data['scale'] = 999
        if 'venues' not in event_data:
            event_data['venues'] = [{}]
        event_data['venues'][0]['name'] = "온라인"
        event_data['venues'][0]['address'] = "온라인"

def handle_budget_info(event_data: Dict[str, Any]) -> None:
    st.header("예산 정보")

    col1, col2 = st.columns(2)
    with col1:
        event_data['contract_status'] = render_option_menu(
            "계약 금액 상태",
            config['CONTRACT_STATUS_OPTIONS'],
            "contract_status"
        )
    with col2:
        vat_options = config['VAT_OPTIONS']
        vat_included = option_menu(
            "부가세 포함 여부",
            options=vat_options,
            icons=['check-circle', 'x-circle'],
            menu_icon="coin",
            default_index=0,
            orientation="horizontal",
            styles={
                "container": {"padding": "0!important", "background-color": "#FFF9C4"},
                "icon": {"color": "#FBC02D", "font-size": "16px"},
                "nav-link": {"font-size": "14px", "text-align": "center", "margin":"0px", "--hover-color": "#FFF59D", "--icon-color": "#FBC02D"},
                "nav-link-selected": {"background-color": "#FBC02D", "color": "white", "--icon-color": "white"},
            },
            key="vat_included"
        )
        event_data['vat_included'] = (vat_included == vat_options[0])

    event_data['contract_amount'] = st.number_input(
        "총 계약 금액 (원)",
        min_value=0,
        value=event_data.get('contract_amount', 0),
        key="contract_amount",
        format="%d"
    )

    display_budget_details(event_data)

    if event_data['contract_status'] == "추가 예정":
        handle_additional_amount(event_data)

    handle_profit_info(event_data)

def display_budget_details(event_data: Dict[str, Any]) -> None:
    if event_data['vat_included']:
        original_amount = round(event_data['contract_amount'] / 1.1)
        vat_amount = round(event_data['contract_amount'] - original_amount)
    else:
        original_amount = event_data['contract_amount']
        vat_amount = round(original_amount * 0.1)

    st.write(f"입력된 계약 금액: {format_currency(event_data['contract_amount'])} 원")
    st.write(f"원금: {format_currency(original_amount)} 원")
    st.write(f"부가세: {format_currency(vat_amount)} 원")

def handle_additional_amount(event_data: Dict[str, Any]) -> None:
    event_data['additional_amount'] = st.number_input(
        "추가 예정 금액 (원)",
        min_value=0,
        value=event_data.get('additional_amount', 0),
        key="additional_amount",
        format="%d"
    )
    st.write(f"입력된 추가 예정 액: {format_currency(event_data['additional_amount'])} 원")

def handle_profit_info(event_data: Dict[str, Any]) -> None:
    event_data['expected_profit_percentage'] = st.number_input(
        "예상 수익률 (%)",
        min_value=0.0,
        max_value=100.0,
        value=event_data.get('expected_profit_percentage', 0.0),
        format="%.2f",
        step=0.01,
        key="expected_profit_percentage"
    )

    total_amount = event_data['contract_amount'] + event_data.get('additional_amount', 0)
    original_amount = round(total_amount / 1.1) if event_data['vat_included'] else total_amount
    expected_profit = round(original_amount * (event_data['expected_profit_percentage'] / 100))

    event_data['expected_profit'] = expected_profit

    st.write(f"예상 수익 금액: {format_currency(expected_profit)} 원")

    total_category_budget = sum(component.get('budget', 0) for component in event_data.get('components', {}).values())
    if total_category_budget > event_data['contract_amount']:
        st.warning(f"주의: 카테고리별 예산 총액({format_currency(total_category_budget)} 원)이 총 계약 금액({format_currency(event_data['contract_amount'])} 원)을 초과합니다.")

def handle_online_content(event_data: Dict[str, Any]) -> None:
    st.subheader("온라인 콘텐츠 정보")
    
    col1, col2 = st.columns(2)
    with col1:
        event_data['start_date'] = st.date_input("콘텐츠 제작 시작일", 
                                   value=event_data.get('start_date', date.today()), 
                                   key="online_start_date")
    with col2:
        event_data['end_date'] = st.date_input("콘텐츠 제작 종료일",
                                 value=max(event_data.get('end_date', event_data['start_date']), event_data['start_date']),
                                 min_value=event_data['start_date'],
                                 key="online_end_date")

    event_data['online_platform'] = st.text_input(
        "사용할 온라인 플랫폼",
        value=event_data.get('online_platform', ''),
        help="예: YouTube, Zoom, Twitch 등",
        key="online_platform"
    )
    
    event_data['streaming_method'] = render_option_menu(
        "스트리밍 방식",
        ["라이브", "녹화 후 업로드", "혼합"],
        "streaming_method"
    )
    
    if event_data['streaming_method'] in ["라이브", "혼합"]:
        event_data['streaming_date'] = st.date_input(
            "스트리밍 날짜",
            value=event_data.get('streaming_date', date.today()),
            key="streaming_date"
        )
        event_data['streaming_time'] = st.time_input(
            "스트리밍 시작 시간",
            value=event_data.get('streaming_time', datetime.now().time()),
            key="streaming_time"
        )
    
    if event_data['streaming_method'] in ["녹화 후 업로드", "혼합"]:
        event_data['recording_location'] = st.text_input(
            "녹화 장소",
            value=event_data.get('recording_location', ''),
            key="recording_location"
        )
        event_data['upload_date'] = st.date_input(
            "업로드 예정일",
            value=event_data.get('upload_date', date.today()),
            key="upload_date"
        )
    
    event_data['expected_duration'] = st.number_input(
        "예상 콘텐츠 길이 (분)",
        min_value=1,
        value=event_data.get('expected_duration', 60),
        key="expected_duration"
    )
    
    event_data['content_description'] = st.text_area(
        "콘텐츠 간단 설명",
        value=event_data.get('content_description', ''),
        key="content_description"
    )

    duration = (event_data['end_date'] - event_data['start_date']).days
    months, days = divmod(duration, 30)
    st.write(f"콘텐츠 제작 기간: {months}개월 {days}일")

def handle_offline_event(event_data: Dict[str, Any]) -> None:
    st.subheader("오프라인 이벤트 정보")

    col1, col2 = st.columns(2)

    with col1:
        start_date = st.date_input("시작 날짜",
                                   value=event_data.get('start_date', date.today()),
                                   key="start_date")

    with col2:
        end_date = st.date_input("종료 날짜",
                                 value=event_data.get('end_date', start_date),
                                 min_value=start_date,
                                 key="end_date")

    event_data['start_date'] = start_date
    event_data['end_date'] = end_date

    col3, col4 = st.columns(2)

    with col3:
        event_data['setup_start'] = render_option_menu("셋업 시작일", config['SETUP_OPTIONS'], "setup_start")

    with col4:
        event_data['teardown'] = render_option_menu("철수 마감일", config['TEARDOWN_OPTIONS'], "teardown")

    if event_data['setup_start'] == config['SETUP_OPTIONS'][0]:
        event_data['setup_date'] = start_date - timedelta(days=1)
    else:
        event_data['setup_date'] = start_date

    if event_data['teardown'] == config['TEARDOWN_OPTIONS'][0]:
        event_data['teardown_date'] = end_date
    else:
        event_data['teardown_date'] = end_date + timedelta(days=1)

    st.write(f"셋업 시작일: {event_data['setup_date']}")
    st.write(f"철수 마감일: {event_data['teardown_date']}")

    if event_data['setup_date'] > start_date:
        st.error("셋업 시작일은 이벤트 시작일보다 늦을 수 없습니다.")
    if end_date < start_date:
        st.error("이벤트 종료일은 시작일보다 빠를 수 없습니다.")
    if event_data['teardown_date'] < end_date:
        st.error("철수 마감일은 이벤트 종료일보다 빠를 수 없습니다.")

def venue_info() -> None:
    event_data = st.session_state.event_data
    st.header("장소 정보")

    guide_text = """
    - **장소 확정 상태**: 현재 장소 섭외 진행 상황을 선택하세요.
    - **희망하는 장소 유형**: 실내, 실외, 혼합, 온라인 중 선택하세요.
    - **예상 참여 관객 수**: 예상되는 참가자 수를 입력하세요.
    - **장소명과 주소**: 확정된 경우 정확한 정보를, 미확정 시 희망 사항을 입력하세요.
    """
    display_guide(guide_text)

    if event_data['event_type'] == "온라인 콘텐츠":
        handle_online_content_location(event_data)
    else:
        handle_offline_event_venue(event_data)

def handle_offline_event_venue(event_data: Dict[str, Any]) -> None:
    event_data['venue_status'] = render_option_menu(
        "장소 확정 상태",
        event_options.STATUS_OPTIONS,
        "venue_status"
    )

    venue_type_options = ["실내", "실외", "혼합", "온라인"]
    event_data['venue_type'] = render_option_menu(
        "희망하는 장소 유형",
        venue_type_options,
        "venue_type"
    )

    if event_data['venue_type'] == "온라인":
        st.info("온라인 이벤트는 물리적 장소 정보가 필요하지 않습니다.")
        event_data['scale'] = 999  # 온라인 이벤트의 경우 scale을 999로 설정
        event_data['venues'] = [{'name': '온라인', 'address': '온라인'}]  # 온라인 이벤트의 경우 장소명과 주소를 '온라인'으로 설정
    else:
        event_data['scale'] = st.number_input(
            "예상 참여 관객 수",
            min_value=0,
            value=event_data.get('scale', 0),
            step=1,
            format="%d",
            key="scale_input_venue"
        )

        if event_data['venue_status'] == "알 수 없는 상태":
            handle_unknown_venue_status(event_data)
        else:
            if 'venues' not in event_data:
                event_data['venues'] = [{'name': '', 'address': ''}]
            handle_known_venue_status(event_data)

def handle_unknown_venue_status(event_data: Dict[str, Any]) -> None:
    major_regions = [
        "서울", "부산", "인천", "대구", "대전", "광주", "울산", "세종",
        "경기도", "강원도", "충청북도", "충청남도", "전라북도", "전라남도", "경상북도", "경상남도", "제주도"
    ]

    def format_region(region: str) -> str:
        region_emojis = {
            "서울": "🗼", "부산": "🌉", "인천": "🛳️", "대구": "🌆", "대전": "🏙️", "광주": "🏞️",
            "울산": "🏭", "세종": "🏛️", "경기도": "🏘️", "강원도": "⛰️", "충청북도": "🌳", "충청남도": "🌊",
            "전라북도": "🍚", "전라남도": "🌴", "경상북도": "🍎", "경상남도": "🐘", "제주도": "🍊"
        }
        return f"{region_emojis.get(region, '📍')} {region}"

    event_data['desired_region'] = st.selectbox(
        "희망하는 지역",
        options=major_regions,
        index=major_regions.index(event_data.get('desired_region', major_regions[0])),
        format_func=format_region,
        key="desired_region_selectbox"
    )

    event_data['specific_location'] = st.text_input("세부 지역 (선택사항)", value=event_data.get('specific_location', ''), key="specific_location")
    event_data['desired_capacity'] = st.number_input("희망하는 수용 인원", min_value=0, value=int(event_data.get('desired_capacity', 0)), key="desired_capacity")

    handle_venue_facilities(event_data)
    handle_venue_budget(event_data)

def handle_known_venue_status(event_data: Dict[str, Any]) -> None:
    if 'venues' not in event_data or not event_data['venues']:
        event_data['venues'] = [{'name': '', 'address': ''}]

    for i, venue in enumerate(event_data['venues']):
        st.subheader(f"장소 {i+1}")
        col1, col2 = st.columns(2)
        with col1:
            venue['name'] = st.text_input("장소명", value=venue.get('name', ''), key=f"venue_name_{i}")
        with col2:
            venue['address'] = st.text_input("주소", value=venue.get('address', ''), key=f"venue_address_{i}")

        if i > 0 and st.button(f"장소 {i+1} 삭제", key=f"delete_venue_{i}"):
            event_data['venues'].pop(i)
            st.experimental_rerun()

    if st.button("장소 추가"):
        event_data['venues'].append({'name': '', 'address': ''})
        st.experimental_rerun()

    handle_venue_facilities(event_data)
    handle_venue_budget(event_data)

def handle_venue_facilities(event_data: Dict[str, Any]) -> None:
    if event_data['venue_type'] in ["실내", "혼합"]:
        facility_options = ["음향 시설", "조명 시설", "LED 시설", "빔프로젝트 시설", "주차", "Wifi", "기타"]
        event_data['facilities'] = st.multiselect("행사장 자체 보유 시설", facility_options, default=event_data.get('facilities', []), key="facilities")

        if "기타" in event_data['facilities']:
            event_data['other_facilities'] = st.text_input("기타 시설 입력", key="other_facility_input")

def handle_venue_budget(event_data: Dict[str, Any]) -> None:
    event_data['venue_budget'] = st.number_input("장소 대관 비용 예산 (원)", min_value=0, value=int(event_data.get('venue_budget', 0)), key="venue_budget", format="%d")

def handle_online_content_location(event_data: Dict[str, Any]) -> None:
    st.subheader("온라인 콘텐츠 정보")
    
    event_data['online_platform'] = st.text_input(
        "사용할 온라인 플랫폼",
        value=event_data.get('online_platform', ''),
        help="예: YouTube, Zoom, Twitch 등",
        key="online_platform"
    )
    
    event_data['streaming_method'] = render_option_menu(
        "스트리밍 방식",
        ["라이브", "녹화 후 업로드", "혼합"],
        "streaming_method"
    )
    
    if event_data['streaming_method'] in ["라이브", "혼합"]:
        event_data['streaming_date'] = st.date_input(
            "스트리밍 날짜",
            value=event_data.get('streaming_date', date.today()),
            key="streaming_date"
        )
        event_data['streaming_time'] = st.time_input(
            "스트리밍 시작 시간",
            value=event_data.get('streaming_time', datetime.now().time()),
            key="streaming_time"
        )
    
    if event_data['streaming_method'] in ["녹화 후 업로드", "혼합"]:
        event_data['recording_location'] = st.text_input(
            "녹화 장소",
            value=event_data.get('recording_location', ''),
            key="recording_location"
        )
        event_data['upload_date'] = st.date_input(
            "업로드 예정일",
            value=event_data.get('upload_date', date.today()),
            key="upload_date"
        )
    
    event_data['expected_duration'] = st.number_input(
        "예상 콘텐츠 길이 (분)",
        min_value=1,
        value=event_data.get('expected_duration', 60),
        key="expected_duration"
    )
    
    event_data['content_description'] = st.text_area(
        "콘텐츠 간단 설명",
        value=event_data.get('content_description', ''),
        key="content_description"
    )

def service_components() -> None:
    event_data = st.session_state.event_data
    st.header("용역 구성 요소")

    selected_categories = select_categories_with_icons(event_data)
    event_data['selected_categories'] = selected_categories

    event_data['components'] = event_data.get('components', {})
    for category in selected_categories:
        handle_category(category, event_data)

    event_data['components'] = {k: v for k, v in event_data['components'].items() if k in selected_categories}

def select_categories_with_icons(event_data: Dict[str, Any]) -> List[str]:
    categories = list(event_options.CATEGORIES.keys())
    default_categories = event_data.get('selected_categories', [])
    default_categories = [cat for cat in default_categories if cat in categories]

    if event_data.get('event_type') == "온라인 콘텐츠" and "미디어" not in default_categories:
        default_categories.append("미디어")
        st.info("온라인 콘텐츠 프로젝트를 위해 '미디어' 카테고리가 자동으로 추가되었습니다.")
    elif event_data.get('venue_type') == "온라인" and "미디어" not in default_categories:
        default_categories.append("미디어")
        st.info("온라인 이벤트를 위해 '미디어' 카테고리가 자동으로 추가되었습니다.")

    col1, col2, col3, col4 = st.columns(4)
    selected_categories = []

    for i, category in enumerate(categories):
        with [col1, col2, col3, col4][i % 4]:
            if st.checkbox(
                f"{event_options.CATEGORY_ICONS[category]} {category}",
                value=category in default_categories,
                key=f"category_{category}_{i}"
            ):
                selected_categories.append(category)

    return selected_categories

def handle_category(category: str, event_data: Dict[str, Any]) -> None:
    st.subheader(category)
    component = event_data['components'].get(category, {})

    component['status'] = render_option_menu(
        f"{category} 진행 상황",
        event_options.STATUS_OPTIONS,
        f"{category}_status"
    )

    component['items'] = st.multiselect(
        f"{category} 항목 선택",
        event_options.CATEGORIES.get(category, []) + ["기타"],
        default=component.get('items', []),
        key=f"{category}_items"
    )

    component['budget'] = st.number_input(f"{category} 예산 (원)", min_value=0, value=component.get('budget', 0), key=f"{category}_budget")

    shooting_date_status = render_option_menu(
        "촬영일이 정해졌나요?",
        ["정해짐", "미정"],
        f"{category}_shooting_date_status"
    )

    st.info("연간 제작건의 경우에도 촬영 시작일과 마감일을 입력해주세요.")

    col1, col2 = st.columns(2)
    with col1:
        component['shooting_start_date'] = st.date_input(
            "촬영 시작일",
            min_value=date.today(),
            key=f"{category}_shooting_start_date"
        )
    with col2:
        component['shooting_end_date'] = st.date_input(
            "촬영 마감일",
            min_value=component['shooting_start_date'],
            key=f"{category}_shooting_end_date"
        )

    component['delivery_dates'] = component.get('delivery_dates', [{}])
    
    for idx, delivery in enumerate(component['delivery_dates']):
        st.subheader(f"납품일 {idx + 1}")
        
        delivery['status'] = render_option_menu(
            "납품일이 정해졌나요?",
            ["정해짐", "미정"],
            f"{category}_delivery_status_{idx}"
        )

        if delivery['status'] == "정해짐":
            delivery_type = render_option_menu(
                "납품 방식을 선택해주세요",
                ["기간", "지정일"],
                f"{category}_delivery_type_{idx}"
            )
            
            if delivery_type == "기간":
                col1, col2 = st.columns(2)
                with col1:
                    delivery['start_date'] = st.date_input(
                        "납품 시작일",
                        min_value=component['shooting_start_date'],
                        key=f"{category}_delivery_start_date_{idx}"
                    )
                with col2:
                    delivery['end_date'] = st.date_input(
                        "납품 마감일",
                        min_value=delivery['start_date'],
                        key=f"{category}_delivery_end_date_{idx}"
                    )
            else:
                delivery['date'] = st.date_input(
                    "납품일을 선택해주세요",
                    min_value=component['shooting_start_date'],
                    key=f"{category}_delivery_date_{idx}"
                )
        else:
            delivery['date'] = None

        delivery['items'] = {}
        for item in component['items']:
            quantity = st.number_input(
                f"{item} 납품 수량",
                min_value=0,
                value=delivery['items'].get(item, 0),
                key=f"{category}_delivery_item_{idx}_{item}"
            )
            if quantity > 0:
                delivery['items'][item] = quantity

    if len(component['delivery_dates']) > 1 and st.button("납품일 삭제", key=f"{category}_remove_delivery_date"):
        component['delivery_dates'].pop()

    if st.button("납품일 추가", key=f"{category}_add_delivery_date"):
        component['delivery_dates'].append({})

    if category == "미디어":
        component['reference_links'] = component.get('reference_links', [''])
        for i, link in enumerate(component['reference_links']):
            component['reference_links'][i] = st.text_input(f"레퍼런스 링크 {i+1} (필수)", value=link, key=f"{category}_reference_link_{i}")
        
        if st.button("레퍼런스 링크 추가", key=f"{category}_add_reference_link"):
            component['reference_links'].append('')

        if len(component['reference_links']) > 1 and st.button("레퍼런스 링크 삭제", key=f"{category}_remove_reference_link"):
            component['reference_links'].pop()

    cooperation_options = ["협력사 매칭 필요", "선호하는 업체 있음"]
    component['cooperation_status'] = render_option_menu(
        "협력사 상태",
        cooperation_options,
        f"{category}_cooperation_status"
    )

    if component['cooperation_status'] == "선호하는 업체 있음":
        handle_preferred_vendor(component, category)
    else:
        component['preferred_vendor'] = False
        component['vendor_reason'] = ''
        component['vendor_name'] = ''
        component['vendor_contact'] = ''
        component['vendor_manager'] = ''

    for item in component['items']:
        if item == "기타":
            handle_other_items(component, category)
        else:
            handle_item_details(item, component)

    total_quantities = {item: 0 for item in component['items']}
    for delivery in component['delivery_dates']:
        for item, quantity in delivery['items'].items():
            total_quantities[item] += quantity

    st.subheader("항목별 총 수량 검토")
    for item in component['items']:
        col1, col2, col3 = st.columns(3)
        with col1:
            st.write(f"{item}:")
        with col2:
            st.write(f"총 납품 수량: {total_quantities[item]}")
        with col3:
            expected_quantity = component.get(f'{item}_quantity', 0)
            st.write(f"예상 수량: {expected_quantity}")
        
        if total_quantities[item] != expected_quantity:
            st.warning(f"{item}의 총 납품 수량과 예상 수량이 일치하지 않습니다.")
        else:
            st.success(f"{item}의 총 납품 수량과 예상 수량이 일치합니다.")

    event_data['components'][category] = component

def handle_preferred_vendor(component: Dict[str, Any], category: str) -> None:
    component['vendor_reason'] = render_option_menu(
        "선호하는 이유를 선택해주세요:",
        config['VENDOR_REASON_OPTIONS'],
        f"{category}_vendor_reason"
    )
    component['vendor_name'] = st.text_input("선호 업체 상호명", value=component.get('vendor_name', ''), key=f"{category}_vendor_name")
    component['vendor_contact'] = st.text_input("선호 업체 연락처", value=component.get('vendor_contact', ''), key=f"{category}_vendor_contact")
    component['vendor_manager'] = st.text_input("선호 업체 담당자명", value=component.get('vendor_manager', ''), key=f"{category}_vendor_manager")

def handle_item_details(item: str, component: Dict[str, Any], item_name: str = None) -> None:
    quantity_key = f'{item}_quantity'
    unit_key = f'{item}_unit'
    duration_key = f'{item}_duration'
    duration_unit_key = f'{item}_duration_unit'
    details_key = f'{item}_details'

    display_name = item_name if item_name else item

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        component[quantity_key] = st.number_input(f"{display_name} 수량", min_value=0, value=component.get(quantity_key, 0), key=quantity_key)

    with col2:
        component[unit_key] = st.text_input(f"{display_name} 단위", value=component.get(unit_key, '개'), key=unit_key)

    with col3:
        component[duration_key] = st.number_input(f"{display_name} 기간", min_value=0, value=component.get(duration_key, 0), key=duration_key)

    with col4:
        component[duration_unit_key] = st.text_input(f"{display_name} 기간 단위", value=component.get(duration_unit_key, '개월'), key=duration_unit_key)

    component[details_key] = st.text_area(f"{display_name} 세부사항", value=component.get(details_key, ''), key=details_key)

def handle_other_items(component: Dict[str, Any], category: str) -> None:
    if 'other_items' not in component:
        component['other_items'] = []

    rerun_needed = False  # rerun이 필요한지 여부를 추적

    for i, other_item in enumerate(component['other_items']):
        col1, col2 = st.columns([3, 1])
        with col1:
            new_value = st.text_input(f"{category} 기타 항목 {i+1}", value=other_item, key=f"{category}_other_item_{i}")
            component['other_items'][i] = new_value
        with col2:
            if st.button(f"삭제 {i+1}", key=f"{category}_delete_other_item_{i}"):
                component['other_items'].pop(i)
                rerun_needed = True  # rerun이 필요함을 표시

    if st.button(f"{category} 기타 항목 추가", key=f"{category}_add_other_item"):
        component['other_items'].append('')
        rerun_needed = True  # rerun이 필요함을 표시

    for i, other_item in enumerate(component['other_items']):
        if other_item:  # 빈 문자열이 아닌 경우에만 처리
            handle_item_details(f"기타_{i+1}", component, item_name=other_item)

    if rerun_needed:
        st.experimental_rerun()  # rerun이 필요한 경우에만 호출

def safe_operation(func):
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            error_msg = f"{func.__name__} 실행 중 오류 발생: {str(e)}"
            st.error(error_msg)
            logging.error(error_msg, exc_info=True)
            return None
    return wrapper

@safe_operation
def generate_summary_excel() -> None:
    event_data = st.session_state.event_data
    event_name = event_data.get('event_name', '무제')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    summary_filename = f"이벤트_기획_정의서_{event_name}_{timestamp}.xlsx"

    try:
        create_excel_summary(event_data, summary_filename)
        st.success(f"엑셀 정의서가 성공적으로 생성되었습니다: {summary_filename}")

        with open(summary_filename, "rb") as file:
            st.download_button(label="전체 행사 요약 정의서 다운로드", data=file, file_name=summary_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        for category, component in event_data.get('components', {}).items():
            category_filename = f"발주요청서_{category}_{event_name}_{timestamp}.xlsx"
            create_category_excel(event_data, category, component, category_filename)
            try:
                with open(category_filename, "rb") as file:
                    st.download_button(label=f"{category} 발주요청서 다운로드", data=file, file_name=category_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"download_{category}")
            except FileNotFoundError:
                st.error(f"{category_filename} 파일을 찾을 수 없습니다.")

    except Exception as e:
        st.error(f"엑셀 파일 생성 중 오류가 발생했습니다: {str(e)}")
        st.error("오류 상세 정보:")
        st.exception(e)

@safe_operation
def create_excel_summary(event_data: Dict[str, Any], filename: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "전체 용역 정의서"

    # 제목
    ws.merge_cells('A1:H1')
    ws['A1'] = '전체 용역 정의서'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 받는 곳
    ws.merge_cells('A3:H3')
    ws['A3'] = '◎ 받는 곳 : ㈜디노마드 / 서울시 영등포구 여의대로 108 파크원타워 2, 21층'
    ws['A3'].alignment = Alignment(horizontal='left')

    # 발주 요청 사항
    ws.merge_cells('A5:H5')
    ws['A5'] = '아래 사항에 대하여 귀사의 견적을 요청하오니 견적서를 제출하여 주시기 바라며,\n견적서 제출 후 계약을 진행하여 주시기 바랍니다.'
    ws['A5'].alignment = Alignment(horizontal='left')

    # 프로젝트 정보
    project_info = [
        ('프로젝트명', event_data.get('event_name', ''), '용역유형', event_data.get('event_type', '')),
        ('고객사', event_data.get('client_name', ''), '담당 PM', f"{event_data.get('manager_name', '')} ({event_data.get('manager_position', '')})"),
        ('담당 PM 연락처', event_data.get('manager_contact', ''), '용역 종류', event_data.get('contract_type', '')),
        ('예상 참여 관객 수', str(event_data.get('scale', '')), '셋업 시작', str(event_data.get('setup_date', ''))),
        ('철수 마감', str(event_data.get('teardown_date', '')), '용역 시작일', str(event_data.get('start_date', ''))),
        ('용역 마감일', str(event_data.get('end_date', '')), '총 계약 금액', f"{format_currency(event_data.get('contract_amount', 0))} 원"),
        ('수익률 / 수익 금액', f"{event_data.get('expected_profit_percentage', 0)}% / {format_currency(event_data.get('expected_profit', 0))} 원", '부가세 포함 여부', '포함' if event_data.get('vat_included', False) else '미포함'),
    ]

    # 이벤트 유형에 따른 추가 정보
    if event_data.get('event_type') == "오프라인 이벤트":
        project_info.extend([
            ('장소', ', '.join([v.get('name', '') for v in event_data.get('venues', [])]), '장소 상태', event_data.get('venue_status', '')),
            ('주소', ', '.join([v.get('address', '') for v in event_data.get('venues', [])]), '', '')
        ])
    elif event_data.get('event_type') == "온라인 콘텐츠":
        project_info.extend([
            ('플랫폼', event_data.get('online_platform', ''), '스트리밍 방식', event_data.get('streaming_method', '')),
            ('촬영 로케이션', event_data.get('location_name', ''), '로케이션 상태', event_data.get('location_status', ''))
        ])

    row = 7
    for item in project_info:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws['A' + str(row)] = item[0]
        ws['A' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws['C' + str(row)] = item[1]
        ws['C' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws['E' + str(row)] = item[2]
        ws['E' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        ws['G' + str(row)] = item[3]
        ws['G' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

        row += 1

    # 구성 요소 헤더
    headers = ['번호', '카테고리', '아이템명', '상세 설명', '수량', '단위', '기간', '기간 단위', '예산', '협력사', '협력사 연락처', '비고']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 아이템 목록
    item_number = 1
    for category, component in event_data.get('components', {}).items():
        for item in component.get('items', []):
            ws.append([
                item_number,
                category,
                item,
                component.get(f'{item}_details', ''),
                component.get(f'{item}_quantity', 0),
                component.get(f'{item}_unit', '개'),
                component.get(f'{item}_duration', 0),
                component.get(f'{item}_duration_unit', '개월'),
                component.get('budget', 0),
                component.get('vendor_name', ''),
                component.get('vendor_contact', ''),
                ''
            ])
            item_number += 1

    # 열 너비 설정
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col].width = 20

    wb.save(filename)

def create_media_summary(event_data: Dict[str, Any], filename: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "미디어 발주 요약"

    # 기본 정보
    ws['A1'] = "미디어 발주 요약서"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:G1')

    basic_info = [
        ('프로젝트명', event_data.get('event_name', '')),
        ('클라이언트', event_data.get('client_name', '')),
        ('담당 PM', event_data.get('manager_name', '')),
        ('연락처', event_data.get('manager_contact', ''))
    ]

    for row, (key, value) in enumerate(basic_info, start=3):
        ws[f'A{row}'] = key
        ws[f'B{row}'] = value

    # 미디어 정보
    media_component = event_data.get('components', {}).get('Media', {})
    
    row = 8
    ws[f'A{row}'] = "촬영 정보"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    if media_component.get('shooting_date'):
        ws[f'A{row}'] = "촬영일"
        ws[f'B{row}'] = str(media_component['shooting_date'])
    else:
        ws[f'A{row}'] = "촬영 기간"
        ws[f'B{row}'] = f"{media_component.get('shooting_start_date', '')} ~ {media_component.get('shooting_end_date', '')}"
    
    row += 2
    ws[f'A{row}'] = "납품 정보"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    for idx, delivery in enumerate(media_component.get('delivery_dates', []), 1):
        ws[f'A{row}'] = f"납품일 {idx}"
        ws[f'B{row}'] = str(delivery['date']) if delivery['date'] else '미정'
        row += 1
        
        ws[f'A{row}'] = "항목"
        ws[f'B{row}'] = "수량"
        row += 1
        
        for item, quantity in delivery['items'].items():
            ws[f'A{row}'] = item
            ws[f'B{row}'] = quantity
            row += 1
        
        row += 1

    # 스타일 적용
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 20

    wb.save(filename)

@safe_operation
def create_category_excel(event_data: Dict[str, Any], category: str, component: Dict[str, Any], filename: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sanitize_sheet_title(category)

    # 제목
    ws.merge_cells('A1:H1')
    ws['A1'] = f'{category} 발주요청서'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # 받는 곳
    ws.merge_cells('A3:H3')
    ws['A3'] = '◎ 받는 곳 : ㈜디노마드 / 서울시 영등포구 여의대로 108 파크원타워 2, 21층'
    ws['A3'].alignment = Alignment(horizontal='left')

    # 발주 요청 사항
    ws.merge_cells('A5:H5')
    ws['A5'] = '아래 사항에 대하여 귀사의 견적을 요청하오니 견적서를 제출하여 주시기 바라며,\n견적서 제출 후 계약을 진행하여 주시기 바랍니다.'
    ws['A5'].alignment = Alignment(horizontal='left')

# 프로젝트 정보
    project_info = [
        ('프로젝트명', event_data.get('event_name', ''), '용역유형', event_data.get('event_type', '')),
        ('고객사', event_data.get('client_name', ''), '담당 PM', f"{event_data.get('manager_name', '')} ({event_data.get('manager_position', '')})"),
        ('담당 PM 연락처', event_data.get('manager_contact', ''), '용역 종류', event_data.get('contract_type', '')),
        ('예상 참여 관객 수', str(event_data.get('scale', '')), '셋업 시작', str(event_data.get('setup_date', ''))),
        ('철수 마감', str(event_data.get('teardown_date', '')), '용역 시작일', str(event_data.get('start_date', ''))),
        ('용역 마감일', str(event_data.get('end_date', '')), '총 계약 금액', f"{format_currency(event_data.get('contract_amount', 0))} 원"),
        ('수익률 / 수익 금액', f"{event_data.get('expected_profit_percentage', 0)}% / {format_currency(event_data.get('expected_profit', 0))} 원", '', ''),
    ]

    row = 7
    for item in project_info:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws['A' + str(row)] = item[0]
        ws['A' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws['C' + str(row)] = item[1]
        ws['C' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws['E' + str(row)] = item[2]
        ws['E' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        ws['G' + str(row)] = item[3]
        ws['G' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

        row += 1

    # 촬영일 정보
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws['A' + str(row)] = '촬영일 정보'
    ws['A' + str(row)].font = Font(bold=True)
    ws['A' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

    row += 1
    if 'shooting_date' in component:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws['A' + str(row)] = f"촬영일: {component['shooting_date']}"
    else:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws['A' + str(row)] = f"촬영 가능 기간: {component['shooting_start_date']} ~ {component['shooting_end_date']}"

    # 납품일 정보
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws['A' + str(row)] = '납품일 정보'
    ws['A' + str(row)].font = Font(bold=True)
    ws['A' + str(row)].alignment = Alignment(horizontal='left', vertical='center')

    for delivery in component.get('delivery_dates', []):
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        if delivery['status'] == "정해짐":
            ws['A' + str(row)] = f"납품일: {delivery['date']}"
        else:
            ws['A' + str(row)] = "납품일: 미정"

        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws['A' + str(row)] = "납품 항목:"
        for item, quantity in delivery['items'].items():
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            ws['A' + str(row)] = f"- {item}: {quantity}개"

    # 아이템 목록
    row += 2
    headers = ['번호', '아이템명', '상세 설명', '수량', '단위', '기간', '기간 단위', '비고']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    item_number = 1
    for item in component.get('items', []):
        row += 1
        ws.append([
            item_number,
            item,
            component.get(f'{item}_details', ''),
            component.get(f'{item}_quantity', 0),
            component.get(f'{item}_unit', '개'),
            component.get(f'{item}_duration', 0),
            component.get(f'{item}_duration_unit', '개월'),
            ''
        ])
        item_number += 1

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].auto_size = True

    wb.save(filename)

def sanitize_sheet_title(title: str) -> str:
    invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
    for char in invalid_chars:
        title = title.replace(char, '')
    return title

def check_required_fields(step):
    event_data = st.session_state.event_data
    missing_fields = []

    if step == 0:  # 기본 정보
        required_fields = ['event_name', 'client_name', 'manager_name', 'manager_contact', 'event_type', 'contract_type']
        for field in required_fields:
            if not event_data.get(field):
                missing_fields.append(field)

    elif step == 1:  # 장소 정보
        if event_data.get('event_type') != "온라인 콘텐츠":
            if not event_data.get('venue_status'):
                missing_fields.append('venue_status')
            if not event_data.get('venue_type'):
                missing_fields.append('venue_type')
            if event_data.get('venue_type') != "온라인":
                if not event_data.get('scale'):
                    missing_fields.append('scale')
                for i, venue in enumerate(event_data.get('venues', [])):
                    if not venue.get('name'):
                        missing_fields.append(f'venues[{i}].name')
                    if not venue.get('address'):
                        missing_fields.append(f'venues[{i}].address')

    elif step == 2:  # 용역 구성 요소
        if not event_data.get('selected_categories'):
            missing_fields.append('selected_categories')
        else:
            for category in event_data.get('selected_categories', []):
                if category not in event_data.get('components', {}):
                    missing_fields.append(f'components.{category}')
                else:
                    component = event_data['components'][category]
                    if not component.get('status'):
                        missing_fields.append(f'components.{category}.status')
                    if not component.get('items'):
                        missing_fields.append(f'components.{category}.items')

    return len(missing_fields) == 0, missing_fields

def highlight_missing_fields(missing_fields):
    field_names = {
        'event_name': '용역명',
        'client_name': '클라이언트명',
        'manager_name': '담당 PM',
        'manager_position': '담당자 직급',
        'manager_contact': '담당자 연락처',
        'event_type': '용역 유형',
        'contract_type': '용역 종류',
        'scale': '예상 참여 관객 수',
        'contract_amount': '총 계약 금액',
        'expected_profit_percentage': '예상 수익률',
        'start_date': '시작일',
        'end_date': '종료일',
        'setup_start': '셋업 시작일',
        'teardown': '철수 마감일',
        'venue_status': '장소 확정 상태',
        'venue_type': '장소 유형',
        'desired_region': '희망하는 지역',
        'desired_capacity': '희망하는 수용 인원',
        'venues': '장소',
        'selected_categories': '선택된 카테고리',
        'components': '용역 구성 요소',
        'status': '상태',
        'items': '항목',
        'name': '장소명',
        'address': '주소',
        'online_platform': '온라인 플랫폼',
        'streaming_method': '스트리밍 방식'
    }

    for field in missing_fields:
        if field == 'invalid_date_range':
            st.error("시작일이 종료일보다 늦을 수 없습니다.")
        elif field == 'invalid_event_dates':
            st.error("이벤트 날짜가 올바르지 않습니다. 셋업 시작일 ≤ 시작일 ≤ 종료일 ≤ 철수 마감일 순서여야 합니다.")
        elif '.' in field:
            category, subfield = field.split('.', 1)
            st.error(f"{field_names.get(category, category)} 카테고리의 {field_names.get(subfield, subfield)} 항목을 입력해주세요.")
        elif '[' in field:
            base, index = field.split('[')
            index = index.split(']')[0]
            subfield = field.split('.')[-1]
            st.error(f"{field_names.get(base, base)} 목록의 {int(index)+1}번째 항목의 {field_names.get(subfield, subfield)}을(를) 입력해주세요.")
        else:
            st.error(f"{field_names.get(field, field)} 항목을 입력해주세요.")

def main():
    st.title("이벤트 플래너")

    if 'current_event' not in st.session_state:
        st.session_state.current_event = None
    if 'step' not in st.session_state:
        st.session_state.step = 0
    if 'event_data' not in st.session_state:
        st.session_state.event_data = {}
    if 'previous_button' not in st.session_state:
        st.session_state.previous_button = False
    if 'next_button' not in st.session_state:
        st.session_state.next_button = False

    functions = {
        0: basic_info,
        1: venue_info,
        2: service_components,
        3: generate_summary_excel
    }

    step_names = ["기본 정보", "장소 정보", "용역 구성 요소", "정의서 생성"]

    current_step = st.session_state.step
    event_type = st.session_state.event_data.get('event_type')

    if event_type == "온라인 콘텐츠" and current_step == 1:
        current_step = 2
        st.session_state.step = 2

    selected_step = option_menu(
        None,
        step_names,
        icons=['info-circle', 'geo-alt', 'list-task', 'file-earmark-spreadsheet'],
        default_index=current_step,
        orientation='horizontal',
        styles={
            "container": {"padding": "0!important", "background-color": "#e3f2fd"},
            "icon": {"color": "#1976d2", "font-size": "25px"},
            "nav-link": {"font-size": "16px", "text-align": "center", "margin":"0px", "--hover-color": "#bbdefb", "--icon-color": "#1976d2"},
            "nav-link-selected": {"background-color": "#2196f3", "color": "white", "--icon-color": "white"},
        },
    )

    if selected_step != step_names[current_step]:
        new_step = step_names.index(selected_step)
        if event_type == "온라인 콘텐츠" and new_step == 1:
            new_step = 2
        st.session_state.step = new_step
        st.experimental_rerun()

    functions[current_step]()

    col1, col2 = st.columns([1, 1])

    if col1.button("이전", key="previous_button") and current_step > 0:
        if event_type == "온라인 콘텐츠" and current_step == 2:
            st.session_state.step = 0
        else:
            st.session_state.step -= 1
        st.experimental_rerun()

    if col2.button("다음", key="next_button") and current_step < 3:
        is_valid, missing_fields = check_required_fields(current_step)
        if is_valid:
            if event_type == "온라인 콘텐츠" and current_step == 0:
                st.session_state.step = 2
            else:
                st.session_state.step += 1
            st.experimental_rerun()
        else:
            highlight_missing_fields(missing_fields)

if __name__ == "__main__":
    main()