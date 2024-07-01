import streamlit as st
import json
import os
import pandas as pd
from datetime import datetime, timedelta, date
from typing import Dict, Any, List
from streamlit_extras.colored_header import colored_header
from streamlit_extras.add_vertical_space import add_vertical_space
from streamlit_extras.metric_cards import style_metric_cards
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import re

# 상수 정의
CONFIRMED = "확정됨"
ALMOST_CONFIRMED = "거의 확정됨"
IN_PROGRESS = "진행 중"
NOT_STARTED = "아직 시작 못함"

# 페이지 구성 설정
st.set_page_config(layout="wide", page_title="이벤트 기획 도구")

# CSS 스타일 적용
st.markdown("""
<style>
    .stApp {
        max-width: 1200px;
        margin: 0 auto;
        font-family: 'Arial', sans-serif;
    }
    .progress-bar {
        width: 100%;
        background-color: #f0f0f0;
        border-radius: 5px;
        margin-bottom: 20px;
    }
    .progress {
        width: 0%;
        height: 20px;
        background-color: #4CAF50;
        border-radius: 5px;
        transition: width 0.5s ease-in-out;
    }
    .step-title {
        font-size: 24px;
        font-weight: bold;
        margin-bottom: 20px;
    }
    .stButton>button {
        border-radius: 5px;
        height: 3em;
        background-color: #4CAF50;
        color: white;
        font-weight: bold;
    }
    .stTextInput>div>div>input {
        border-radius: 5px;
    }
    h1, h2, h3 {
        color: #2c3e50;
    }
</style>
""", unsafe_allow_html=True)

def format_korean_currency(amount):
    if amount == "미정":
        return amount
    try:
        amount = int(amount)
        return '{:,}'.format(amount)
    except ValueError:
        return amount

def parse_korean_currency(amount):
    if amount == "미정":
        return amount
    try:
        return int(amount.replace(',', ''))
    except ValueError:
        return amount

def budget_input(key, label):
    col1, col2 = st.columns([3, 1])
    with col1:
        amount = st.text_input(label, value=format_korean_currency(st.session_state.data.get(key, '')))
    with col2:
        is_undecided = st.checkbox("미정", key=f"{key}_undecided")
    
    if is_undecided:
        st.session_state.data[key] = "미정"
    else:
        st.session_state.data[key] = parse_korean_currency(amount)

def main():
    if 'step' not in st.session_state:
        st.session_state.step = 1
    if 'data' not in st.session_state:
        st.session_state.data = {}

    steps = ["기본 정보", "용역 개요", "용역 형태 및 장소", "용역 구성 요소", "마무리"]
    total_steps = len(steps)

    display_progress(st.session_state.step, total_steps)

    colored_header(
        label=f"Step {st.session_state.step}: {steps[st.session_state.step - 1]}",
        description="용역 기획을 위한 단계별 가이드",
        color_name="blue-70"
    )

    if st.session_state.step == 1:
        display_basic_info()
    elif st.session_state.step == 2:
        display_service_overview()
    elif st.session_state.step == 3:
        display_service_format_and_venue()
    elif st.session_state.step == 4:
        display_service_components()
    elif st.session_state.step == 5:
        finalize()

    col1, col2 = st.columns(2)
    with col1:
        if st.session_state.step > 1:
            if st.button("이전", key="prev_button"):
                st.session_state.step -= 1
                st.experimental_rerun()
    with col2:
        if st.session_state.step < total_steps:
            if st.button("다음", key="next_button"):
                if validate_current_step():
                    st.session_state.step += 1
                    st.experimental_rerun()
                else:
                    st.error("모든 필수 항목을 채워주세요.")

    add_vertical_space(2)

def display_progress(current_step: int, total_steps: int):
    progress = (current_step / total_steps) * 100
    st.markdown(f"""
        <div class="progress-bar">
            <div class="progress" style="width: {progress}%;"></div>
        </div>
        <p style="text-align: center;">진행 상황: {current_step} / {total_steps}</p>
    """, unsafe_allow_html=True)

def improved_schedule_input():
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("용역 시작일", value=st.session_state.data.get('service_start_date', datetime.now().date()))
    
    if 'service_end_date' not in st.session_state.data or st.session_state.data['service_end_date'] < start_date:
        end_date = start_date
    else:
        end_date = st.session_state.data['service_end_date']
    
    with col2:
        end_date = st.date_input("용역 종료일", value=end_date, min_value=start_date)
    
    st.session_state.data['service_start_date'] = start_date
    st.session_state.data['service_end_date'] = end_date

    setup_options = ["전날부터", "당일"]
    teardown_options = ["당일 철수", "다음날 철수"]

    col3, col4 = st.columns(2)
    with col3:
        setup_choice = st.radio("셋업 시작", setup_options, index=setup_options.index(st.session_state.data.get('setup_choice', '전날부터')))
    with col4:
        teardown_choice = st.radio("철수", teardown_options, index=teardown_options.index(st.session_state.data.get('teardown_choice', '당일 철수')))

    st.session_state.data['setup_choice'] = setup_choice
    st.session_state.data['teardown_choice'] = teardown_choice

    st.session_state.data['setup_date'] = start_date - timedelta(days=1) if setup_choice == "전날부터" else start_date
    st.session_state.data['teardown_date'] = end_date if teardown_choice == "당일 철수" else end_date + timedelta(days=1)

    st.write(f"셋업 일자: {st.session_state.data['setup_date']}")
    st.write(f"철수 일자: {st.session_state.data['teardown_date']}")

    service_duration = (end_date - start_date).days + 1
    st.session_state.data['service_duration'] = service_duration
    st.write(f"용역 기간: {service_duration}일")

def display_basic_info():
    st.session_state.data['name'] = st.text_input("이름", st.session_state.data.get('name', ''))
    st.session_state.data['department'] = st.text_input("근무 부서", st.session_state.data.get('department', ''))
    
    position_options = ["파트너 기획자", "선임", "책임", "수석"]
    st.session_state.data['position'] = st.radio("직급", position_options, index=position_options.index(st.session_state.data.get('position', '파트너 기획자')))
    
    service_types = ["행사 운영", "공간 디자인", "마케팅", "PR", "영상제작", "전시", "브랜딩", "온라인 플랫폼 구축", "기타"]
    st.session_state.data['service_types'] = st.multiselect("주로 하는 용역 유형", service_types, default=st.session_state.data.get('service_types', []))
    
    st.session_state.data['service_name'] = st.text_input("용역명", st.session_state.data.get('service_name', ''))

    improved_schedule_input()

def display_service_overview():
    service_purposes = ["브랜드 인지도 향상", "고객 관계 강화", "신제품 출시", "교육 및 정보 제공", "수익 창출", "문화/예술 증진", "기타"]
    st.session_state.data['service_purpose'] = st.multiselect("용역의 주요 목적", service_purposes, default=st.session_state.data.get('service_purpose', []))
    
    # 예상 참가자 수 입력 방식 선택
    input_method = st.radio("예상 참가자 수 입력 방식", ["단일 값", "범위 설정"], index=0)
    
    if input_method == "단일 값":
        col1, col2 = st.columns(2)
        with col1:
            use_slider = st.checkbox("슬라이더 사용", value=False)
        with col2:
            participants_over_2000 = st.checkbox("2000명 이상", value=st.session_state.data.get('participants_over_2000', False))
        
        if participants_over_2000:
            st.session_state.data['expected_participants'] = "2000명 이상"
        else:
            if use_slider:
                current_value = st.session_state.data.get('expected_participants', 100)
                if isinstance(current_value, str):
                    try:
                        current_value = int(current_value)
                    except ValueError:
                        current_value = 100
                st.session_state.data['expected_participants'] = st.slider("예상 참가자 수", 0, 2000, value=int(current_value))
            else:
                current_value = st.session_state.data.get('expected_participants', '')
                if isinstance(current_value, int):
                    current_value = str(current_value)
                st.session_state.data['expected_participants'] = st.text_input("예상 참가자 수", value=current_value)
    else:
        col1, col2 = st.columns(2)
        with col1:
            min_participants = st.number_input("최소 예상 참가자 수", value=st.session_state.data.get('min_participants', 0), min_value=0)
        with col2:
            max_participants = st.number_input("최대 예상 참가자 수", value=st.session_state.data.get('max_participants', 100), min_value=min_participants)
        
        st.session_state.data['min_participants'] = min_participants
        st.session_state.data['max_participants'] = max_participants
        st.session_state.data['expected_participants'] = f"{min_participants} - {max_participants}"

    contract_types = ["수의계약", "입찰", "B2B"]
    st.session_state.data['contract_type'] = st.radio("계약 형태", contract_types, index=contract_types.index(st.session_state.data.get('contract_type', '수의계약')))
    
    budget_statuses = [CONFIRMED, ALMOST_CONFIRMED, IN_PROGRESS, NOT_STARTED]
    st.session_state.data['budget_status'] = st.radio("예산 협의 상태", budget_statuses, index=budget_statuses.index(st.session_state.data.get('budget_status', IN_PROGRESS)))

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric(label="예상 참가자 수", value=st.session_state.data['expected_participants'])
    with col2:
        st.metric(label="계약 형태", value=st.session_state.data['contract_type'])
    with col3:
        st.metric(label="예산 협의 상태", value=st.session_state.data['budget_status'])
    
    style_metric_cards()

def display_service_format_and_venue():
    service_formats = ["오프라인", "온라인", "하이브리드", "기타"]
    st.session_state.data['service_format'] = st.radio("용역 형태", service_formats, index=service_formats.index(st.session_state.data.get('service_format', '오프라인')))
    
    if st.session_state.data['service_format'] in ["오프라인", "하이브리드"]:
        venue_types = ["실내 (호텔, 컨벤션 센터 등)", "야외 (공원, 광장 등)", "혼합형 (실내+야외)", "아직 미정"]
        st.session_state.data['venue_type'] = st.radio("용역 장소 유형", venue_types, index=venue_types.index(st.session_state.data.get('venue_type', '실내 (호텔, 컨벤션 센터 등)')))
        
        venue_statuses = [CONFIRMED, ALMOST_CONFIRMED, IN_PROGRESS, NOT_STARTED]
        st.session_state.data['venue_status'] = st.radio("용역 장소 협의 상태", venue_statuses, index=venue_statuses.index(st.session_state.data.get('venue_status', IN_PROGRESS)))
        
        if st.session_state.data['venue_status'] in [CONFIRMED, ALMOST_CONFIRMED]:
            st.session_state.data['specific_venue'] = st.text_input("구체적인 장소", st.session_state.data.get('specific_venue', ''))
        else:
            st.session_state.data['expected_area'] = st.text_input("예정 지역", st.session_state.data.get('expected_area', ''))

def display_service_components():
    st.header("용역 구성 요소")
    
    components = {
        "무대 설치": setup_stage,
        "음향 시스템": setup_sound,
        "조명 장비": setup_lighting,
        "LED 스크린": setup_led,
        "동시통역 시스템": setup_interpretation,
        "케이터링 서비스": setup_catering,
        "영상 제작": setup_video_production,
        "마케팅 및 홍보": setup_marketing_and_promotion,
        "인력 관리": setup_staff_management,
        "기술 및 장비": setup_technology_and_equipment,
        "네트워킹": setup_networking,
        "예산 및 스폰서십": setup_budget_and_sponsorship,
        "리스크 관리": setup_risk_management,
        "PR": setup_pr,
        "브랜딩": setup_branding,
        "온라인 플랫폼": setup_online_platform
    }
    
    for component, setup_func in components.items():
        with st.expander(f"{component}"):
            setup_func()

def setup_component(component: str, options: Dict[str, Any]):
    data = st.session_state.data.get(component, {})
    data['needed'] = st.checkbox(f"{component} 필요", data.get('needed', False))
    if data['needed']:
        for key, value in options.items():
            if isinstance(value, list):
                data[key] = st.multiselect(key, value, default=data.get(key, []))
            elif isinstance(value, bool):
                data[key] = st.checkbox(key, data.get(key, False))
            elif isinstance(value, int):
                data[key] = st.number_input(key, min_value=value, value=data.get(key, value))
            else:
                data[key] = st.text_input(key, data.get(key, ''))
        
        budget_input(f"{component}_budget", f"{component} 예산")

    st.session_state.data[component] = data

def setup_stage():
    options = {
        "무대에 필요한 기능": ["발표/연설", "음악 공연", "댄스 퍼포먼스", "연극", "시상식", "기타"],
        "무대 형태": ["단상형 (ㅡ 모양)", "일자형 (ㅣ 모양)", "T자형", "기타"],
        "무대 크기": ""
    }
    setup_component("무대 설치", options)

def setup_sound():
    options = {
        "음향 시스템 설치 방식": ["공간과 관객수에 맞게 알아서", "트러스 설치", "바닥 스탠딩", "미정"],
        "장소 답사 필요": True,
        "행사 하우스 음악 필요": True,
        "기념식/시상식 BGM 필요": True
    }
    setup_component("음향 시스템", options)

def setup_lighting():
    options = {
        "필요한 조명 장비": ["일반 조명", "무대 조명", "특수 조명", "LED 조명", "이동식 조명", "기타"]
    }
    setup_component("조명 장비", options)

def setup_led():
    options = {
        "스크린 사용 목적": ["프레젠테이션 표시", "라이브 중계", "배경 그래픽", "무대 세트", "기타"],
        "스크린 크기": ["무대에 꽉 차게", "무대보다 약간 작게", "용도에 맞게 알아서", "정확한 규격 입력"],
        "LED 스크린 크기": ""
    }
    setup_component("LED 스크린", options)

def setup_interpretation():
    options = {
        "필요한 언어": ["영어", "중국어", "일본어", "스페인어", "프랑스어", "기타"],
        "필요한 통역 부스 수": 1,
        "통역 장비 대여 필요": True
    }
    setup_component("동시통역 시스템", options)

def setup_catering():
    options = {
        "식사 유형": ["뷔페", "플레이티드 서비스", "칵테일 리셉션", "도시락", "기타"],
        "스탭 식사 수": 0,
        "VIP/클라이언트 식사 수": 0,
        "특별 식단 요구사항": ""
    }
    setup_component("케이터링 서비스", options)

def setup_video_production():
    options = {
        "필요한 영상 유형": ["행사 홍보 영상", "행사 중계 영상", "하이라이트 영상", "인터뷰 영상", "기타"],
        "예상 영상 길이": "",
        "영상 목적 및 활용 계획": "",
        "원하는 영상 스타일 또는 참고 영상": "",
        "영상 장비 대여 필요": True,
        "필요한 영상 장비": ["카메라", "조명", "마이크", "드론", "기타"],
        "영상 편집 필요": True,
        "영상 편집 요구사항": ""
    }
    setup_component("영상 제작", options)

def setup_marketing_and_promotion():
    options = {
        "사용할 홍보 채널": ["소셜 미디어", "이메일 마케팅", "언론 보도", "온라인 광고", "오프라인 광고 (현수막, 포스터 등)", "인플루언서 마케팅", "기타"],
        "행사 브랜딩 계획 상태": [CONFIRMED, ALMOST_CONFIRMED, IN_PROGRESS, NOT_STARTED],
        "브랜딩 계획 설명": ""
    }
    setup_component("마케팅 및 홍보", options)

def setup_staff_management():
    options = {
        "행사 당일 필요한 예상 인력 수": "",
        "외부에서 고용할 필요가 있는 인력": ["행사 진행 요원", "보안 요원", "기술 지원 인력", "MC 또는 사회자", "공연자 또는 연사", "촬영 및 영상 제작 팀", "기타"],
        "공연자 또는 연사에 대한 추가 정보": ""
    }
    setup_component("인력 관리", options)

def setup_technology_and_equipment():
    options = {
        "필요한 기술적 요구사항": ["와이파이 네트워크", "라이브 스트리밍 장비", "촬영 장비", "행사 관리 소프트웨어", "모바일 앱", "VR/AR 기술", "기타"],
        "기술적 요구사항 세부 설명": "",
        "기타 기술적 요구사항": "",
        "장비 대여 필요": True,
        "대여 필요 장비 목록": ""
    }
    setup_component("기술 및 장비", options)

def setup_networking():
    options = {
        "네트워킹 장소 유형": ["실내", "야외", "기타"],
        "네트워킹 예상 인원 수": 1,
        "네트워킹에 필요한 장비": ["테이블", "의자", "음향 장비", "조명", "기타"],
        "계획된 네트워킹 활동": ""
    }
    setup_component("네트워킹", options)

def setup_budget_and_sponsorship():
    options = {
        "총 예산": "",
        "예산 항목별 내역": "",
        "스폰서 유치 계획 상태": [CONFIRMED, ALMOST_CONFIRMED, IN_PROGRESS, NOT_STARTED],
        "고려 중인 스폰서십 유형": ["금전적 후원", "현물 협찬", "미디어 파트너십", "기술 파트너십", "기타"],
        "스폰서십 세부 사항": ""
    }
    setup_component("예산 및 스폰서십", options)

def setup_risk_management():
    options = {
        "우려되는 잠재적 리스크": ["날씨 (야외 행사의 경우)", "예산 초과", "참가자 수 미달", "기술적 문제", "안전 및 보안 문제", "기타"],
        "리스크 대비책 수립 상태": ["완료", "진행 중", "시작 전", "도움 필요"],
        "리스크 대비책 설명": "",
        "행사 보험 필요 여부": True,
        "필요한 보험 유형": ["행사 취소 보험", "책임 보험", "재산 보험", "기타"]
    }
    setup_component("리스크 관리", options)

def setup_pr():
    options = {
        "PR 목표": ["미디어 노출 증대", "브랜드 이미지 개선", "위기 관리", "기타"],
        "타겟 미디어": ["신문", "방송", "온라인 뉴스", "전문 매체", "기타"],
        "보도자료 작성 필요": True,
        "인터뷰 주선 필요": True,
        "PR 전략 설명": ""
    }
    setup_component("PR", options)

def setup_branding():
    options = {
        "브랜딩 목표": ["브랜드 인지도 향상", "브랜드 이미지 개선", "신규 브랜드 런칭", "기타"],
        "브랜딩 요소": ["로고", "슬로건", "컬러 팔레트", "폰트", "기타"],
        "브랜드 가이드라인 존재 여부": True,
        "브랜드 메시지": "",
        "브랜딩 전략 설명": ""
    }
    setup_component("브랜딩", options)

def setup_online_platform():
    options = {
        "플랫폼 유형": ["웹사이트", "모바일 앱", "가상 이벤트 플랫폼", "기타"],
        "필요한 기능": ["참가자 등록", "라이브 스트리밍", "채팅", "네트워킹", "콘텐츠 공유", "기타"],
        "예상 동시 접속자 수": 0,
        "보안 요구사항": "",
        "플랫폼 개발 기간": "",
        "플랫폼 유지보수 계획": ""
    }
    setup_component("온라인 플랫폼", options)

def generate_excel_file(data: Dict[str, Any], category: str) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{category} 발주요청서"

    # 스타일 정의
    header_font = Font(name='맑은 고딕', size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # 헤더 작성
    headers = ['항목', '내용']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # 공통 정보 작성
    common_info = [
        ('이름', data['name']),
        ('근무 부서', data['department']),
        ('직급', data['position']),
        ('주로 하는 용역 유형', ', '.join(data['service_types'])),
        ('용역명', data['service_name']),
        ('용역 시작일', data['service_start_date']),
        ('용역 마감일', data['service_end_date']),
        ('진행 일정 (일 수)', data['service_duration']),
        ('셋업 시작일', data['setup_date']),
        ('철수 마감일', data['teardown_date']),
        ('용역 목적', ', '.join(data['service_purpose'])),
        ('예상 참가자 수', data['expected_participants']),
        ('계약 형태', data['contract_type']),
        ('예산 협의 상태', data['budget_status']),
        ('용역 형태', data['service_format']),
        ('용역 장소 유형', data.get('venue_type', 'N/A')),
        ('용역 장소', data.get('specific_venue', data.get('expected_area', 'N/A')))
    ]

    for row, (item, content) in enumerate(common_info, start=2):
        ws.cell(row=row, column=1, value=item).border = border
        ws.cell(row=row, column=2, value=content).border = border

    # 카테고리별 정보 작성
    if category in data:
        ws.append(['', ''])  # 빈 행 추가
        ws.append([f'{category} 세부 정보', ''])
        for item, content in data[category].items():
            ws.append([item, str(content)])
            ws.cell(row=ws.max_row, column=1).border = border
            ws.cell(row=ws.max_row, column=2).border = border

    # 예산 정보 추가
    if f"{category}_budget" in data:
        ws.append(['', ''])
        ws.append([f'{category} 예산', format_korean_currency(data[f"{category}_budget"])])
        ws.cell(row=ws.max_row, column=1).border = border
        ws.cell(row=ws.max_row, column=2).border = border

    # 열 너비 조정
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

    # 메모리에 엑셀 파일 저장
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file

def save_survey_data(data: Dict[str, Any]) -> bool:
    try:
        # 저장 디렉토리 생성
        save_dir = "survey_results"
        os.makedirs(save_dir, exist_ok=True)
        
        # 파일명 생성 (이름_날짜_시간.json)
        filename = f"{data['name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(save_dir, filename)
        
        # 데이터 전처리
        processed_data = data.copy()
        for key, value in processed_data.items():
            if isinstance(value, (datetime, date)):
                processed_data[key] = value.isoformat()
        
        # JSON 파일로 저장
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(processed_data, f, ensure_ascii=False, indent=4)
        
        return True
    except Exception as e:
        st.error(f"데이터 저장 중 오류 발생: {str(e)}")
        return False

def finalize():
    st.header("설문 완료")
    st.write("수고하셨습니다! 모든 단계를 완료하셨습니다.")
    
    if st.button("설문 저장 및 발주요청서 다운로드"):
        # 총 예산 계산
        total_budget = 0
        for category in ["무대 설치", "음향 시스템", "조명 장비", "LED 스크린", "동시통역 시스템", "케이터링 서비스", "영상 제작", "마케팅 및 홍보", "PR", "브랜딩", "온라인 플랫폼"]:
            budget_key = f"{category}_budget"
            if budget_key in st.session_state.data:
                budget = st.session_state.data[budget_key]
                if isinstance(budget, (int, float)):
                    total_budget += budget

        st.write(f"총 예산: {format_korean_currency(total_budget)}원")

        # 설문 데이터 저장
        save_result = save_survey_data(st.session_state.data)
        if save_result:
            st.success("설문이 성공적으로 저장되었습니다.")
        else:
            st.error("설문 저장 중 오류가 발생했습니다.")
        
        # 발주요청서 생성 및 다운로드
        categories = ["무대 설치", "음향 시스템", "조명 장비", "LED 스크린", "동시통역 시스템", "케이터링 서비스", "영상 제작", "마케팅 및 홍보", "PR", "브랜딩", "온라인 플랫폼"]
        for category in categories:
            if category in st.session_state.data and st.session_state.data[category].get('needed', False):
                excel_file = generate_excel_file(st.session_state.data, category)
                st.download_button(
                    label=f"{category} 발주요청서 다운로드",
                    data=excel_file,
                    file_name=f"{category}_발주요청서_{st.session_state.data['name']}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        st.success("발주요청서가 생성되었습니다. 위의 버튼을 클릭하여 각 카테고리별 발주요청서를 다운로드 하세요.")

def validate_current_step() -> bool:
    step = st.session_state.step
    data = st.session_state.data

    if step == 1:
        required_fields = ['name', 'department', 'position', 'service_types', 'service_name', 'service_start_date', 'service_end_date']
    elif step == 2:
        required_fields = ['service_purpose', 'expected_participants', 'contract_type', 'budget_status']
    elif step == 3:
        required_fields = ['service_format']
        if data['service_format'] in ["오프라인", "하이브리드"]:
            required_fields.extend(['venue_type', 'venue_status'])
    elif step == 4:
        # 여기서는 모든 필드가 옵션이므로 항상 True를 반환
        return True
    else:
        return True

    for field in required_fields:
        if field not in data or not data[field]:
            return False
    return True

if __name__ == "__main__":
    main()
